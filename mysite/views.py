from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.http import HttpResponse, StreamingHttpResponse, HttpResponseRedirect
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

import os

import openpyxl

from account.models import UserInfo

@csrf_exempt
@require_POST
def get_model(request):
    model = request.POST['model']
    if model == "userlist":
        filename = '用户信息-模板.xlsx'
        path = os.getcwd() + os.sep + 'models'
        file = os.path.join(path, filename)
        if os.path.exists(file):
            return HttpResponse(file.encode('utf-8'))
        else:
            return HttpResponse("error")

def download_file(request):
    if request.GET.get('url', '') != '':
        print('有参数啦')
        filename = request.GET["url"]
        # file_name = filename.split(os.sep)[-1]
        print(filename)
        file_name = filename.split(os.sep)[-1]
        def file_iterator(file_name, chunk_size=512):
            with open(file_name, 'rb') as f:
                while True:
                    c = f.read(chunk_size)
                    if c:
                        yield c
                    else:
                        break

        response = StreamingHttpResponse(file_iterator(filename))
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename=' + file_name.encode('utf-8').decode('ISO-8859-1')
        return response
    return HttpResponse('error')

# @csrf_exempt
@require_POST
def upload_file(request):
    if request.FILES.get('file', '') != '':
        file_obj = request.FILES.get('file')
        # pageIndex = request.GET.get('pageIndex', '1')
        path = os.getcwd() + os.sep + 'models'
        savename = '上传-用户名单.xlsx'
        filepath = os.path.join(path, savename)
        dest = open(filepath, 'wb+')
        dest.write(file_obj.read())
        dest.close()

        wb = openpyxl.load_workbook(filepath)
        sheet_names = wb.get_sheet_names()
        if 'Sheet1' in sheet_names:
            ws = wb.get_sheet_by_name('Sheet1')
            # 检查该条记录是否已经存在
            max_row = ws.max_row
            print(max_row)
            for index in range(2, max_row + 1):
                username = str(ws['A' + str(index)].value).strip()
                row_password = 'szu'+username[-6:]
                password = make_password(row_password)
                name = ws['B' + str(index)].value.strip()
                sex = ws['C' + str(index)].value.strip()
                major = ws['D' + str(index)].value.strip()
                institute = str(ws['E' + str(index)].value).strip()
                classID = str(ws['F' + str(index)].value).strip()
                user_obj = User.objects.filter(username=username)
                if user_obj:
                    pass
                else:
                    User.objects.create(username=username, password=password)
                user_instance = User.objects.get(username=username)
                userinfo_obj = UserInfo.objects.filter(user=user_instance)
                if userinfo_obj:
                    pass
                else:
                    UserInfo.objects.create(user=user_instance, name=name, sex=sex, major=major,
                                         institute=institute, classID=classID)

        else:
            return HttpResponse('上传文件格式有误！')
        wb.save(filepath)
        return HttpResponse('上传成功！')
    return HttpResponseRedirect('无文件上传！')
